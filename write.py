from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook("/Users/joshkrebs/clocky/xlsx_to_xml/AmazonReviews_June2022 (3).xlsx")

ws = wb.worksheets[0]

doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'

xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'


# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)

with tag('reviews'):
    for row in ws.iter_rows(min_row=2, max_row=914, min_col=1, max_col=10):
        row = [cell.value for cell in row]
        with tag("review"):
            with tag("review_id"):
                text(str(row[0]))
            with tag("title"):
                text(str(row[1]))
            with tag("name"):
                text(str(row[2]))
            with tag("review_url"):
                text(str(row[3]))
            with tag("ratings"):
                with tag("overall"):
                  text(row[4])
            with tag("pros"):
                text(str(row[6]))
            if (row[9]):
              with tag("reviewer_image"):
                  text(str(row[9]))

result = indent(
    doc.getvalue(),
    indentation='   ',
    indent_text=True
)

with open("output.xml", "w") as f:
    f.write(result)