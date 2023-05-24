from openpyxl import load_workbook

wb = load_workbook("/Users/joshkrebs/clocky/xlsx_to_xml/AmazonReviews_June2022 (3).xlsx")

ws = wb.worksheets[0]

for row in ws.iter_rows(min_row = 1, max_row= 2, min_col = 1, max_col = 10):
    print([cell.value for cell in row])

